import discord
import os
import asyncio
from discord.ext import commands
from bs4 import BeautifulSoup
import requests
import openpyxl as xl
import pandas as pd
import random
import time
import re
import warnings
import urllib
from urllib.request import URLError
from urllib.request import HTTPError
from urllib.request import urlopen
from urllib.request import Request, urlopen
from urllib.parse import quote

# 오늘의 운세 크롤링 후 엑셀파일로 저장
def crawl_daily_fortune():
    url = 'https://unse.daily.co.kr/?p=zodiac'
    res = requests.get(url)
    soup = BeautifulSoup(res.text, 'html.parser')

    parent = soup.select_one('#card')
    children = parent.select('ul')      # select: 여러개의 요소 / select_one: 가장 첫번째 요소

    wb = xl.Workbook()
    children = parent.select('ul')      # select: 여러개의 요소 / select_one: 가장 첫번째 요소
    for i in children:
        lis = i.select('li')
        for n, li in enumerate(lis):
            if n == 0:
                animal = li.select_one('div > b').get_text()    # 태그 중 텍스트 추출
                total = li.select_one('div > p').get_text()
                ws = wb.create_sheet(animal+'띠')    # animal 이름으로 시트 생성
                ws.append([total])    
            else:
                year = li.select_one('span').get_text()
                luck = li.select_one('p').get_text()
                ws.append([year, luck])
    wb.save('fortune.xlsx')

def movie_rank():
    boxoffice = "https://movie.daum.net/ranking/boxoffice/weekly"
    
    res1 = requests.get(boxoffice)
    res1.raise_for_status()
    bs1  = BeautifulSoup(res1.text, "lxml")

    parent1 = bs1.select_one('#mainContent')
    movies = parent1.find_all("div", attrs={"class":"item_poster"})
    
    wb = xl.Workbook()
    ws_rank = wb.create_sheet("영화 순위")
    ws_info = wb.create_sheet("영화 정보")
    ws_rank.append(['랭크', '타이틀', '스토리', '링크', '포스터'])

    for movie in movies:
        movie_rank = movie.find("span").get_text()
        movie_title = movie.find("strong").get_text()
        movie_link =  "https://movie.daum.net" + movie.a["href"]
        movie_story = movie.a.get_text()
        movie_poster = movie.find("img")
        movie_poster_src = movie_poster["src"]
        ws_rank.append(["{}위.".format(movie_rank), movie_title, movie_story, movie_link, movie_poster_src])
    wb.save('MovieRank.xlsx')

def movie_info(num_url):
    url_movie = num_url

    res2 = requests.get(url_movie)
    res2.raise_for_status()
    bs2  = BeautifulSoup(res2.text, "lxml")

    parent2 = bs2.select_one('#mainContent')
    s_movie_infos = parent2.find_all("dl", attrs={"class":"list_cont"})

    wb = xl.load_workbook('C:/discordbot/MovieRank.xlsx')
    ws_info = wb["영화 정보"]

    for s_movie in s_movie_infos:
        s_movie_list = s_movie.find("dd").get_text()
        s_movie_info = s_movie.find("dt").get_text()
        ws_info.append([s_movie_info, s_movie_list])
    wb.save('MovieRank.xlsx')


# 디스코드 봇 실행 코드

game = discord.Game("ㄹㅇ 도움요청")
bot = commands.Bot(command_prefix='ㄹㅇ ', status=discord.Status.online, activity=game)

@bot.event
async def on_ready():
    print('사용중인 봇 :', bot.user.name)
    print('정상적으로 작동중입니다.\n')
    await bot.change_presence(status=discord.Status.online, activity=game)

@bot.command(aliases=['안녕', 'hi', '안녕하세요', 'ㅎㅇ'])
async def hello(ctx):
    random_num = random.randrange(1,15)
    if random_num >= 4:
        await ctx.send(f'{ctx.author.mention}님 안녕하세요!')
    elif random_num < 4: # 낮은 확률로 다른 인사
        await ctx.send(f'어이 {ctx.author.mention} 드디어 왔구만!')

@bot.command()
async def 도움요청(ctx):
    embed = discord.Embed(title='도움말', description='제작 by Seung_hoen', color = 0x00ff00)
    embed.add_field(name='1. 인사', value='ex) ㄹㅇ 안녕', inline=False)
    embed.add_field(name='2. 오늘의 운세', value='ex) ㄹㅇ 운세 <~띠>', inline=False)
    embed.add_field(name='3. 국내 코로나 현황', value='ex) ㄹㅇ 코로나', inline=False)
    embed.add_field(name='4. 이번주 영화 순위', value='ex) ㄹㅇ 영화', inline=False)
    embed.add_field(name='5. 순위권 영화 정보', value='ex) ㄹㅇ 영화정보 <순위>', inline=False)
    await ctx.send(embed=embed)

@bot.command()
async def 운세(ctx, *, text):
    crawl_daily_fortune() # 크롤링
    wb = xl.load_workbook('파일경로/fortune.xlsx') 
    sheet = wb[text]

    one_line = [sheet['A1'].value]
    year_line = []

    for data in sheet['A4':'B6']:
        for cell in data:
            year_line.append(cell.value)
    wb.close()

    embed = discord.Embed(title=text, description=one_line[0], color=0xffff00)
    embed.add_field(name=year_line[0], value=year_line[1], inline=False)
    embed.add_field(name=year_line[2], value=year_line[3], inline=False)
    embed.add_field(name=year_line[4], value=year_line[5], inline=False)
    embed.add_field(name='더 보기...', value='Link : https://unse.daily.co.kr/?p=zodiac', inline=False)
    await ctx.send(embed=embed)

@bot.command()
async def 코로나(ctx):
    # 보건복지부 코로나 바이러스 정보사이트"
    covidSite = "http://ncov.mohw.go.kr/index.jsp"
    covidNotice = "http://ncov.mohw.go.kr"
    html = urlopen(covidSite)
    bs = BeautifulSoup(html, 'html.parser')
    latestupdateTime = bs.find('span', {'class': "livedate"}).text.split(',')[0][1:].split('.')
    statisticalNumbers = bs.findAll('span', {'class': 'num'})
    beforedayNumbers = bs.findAll('span', {'class': 'before'})

    #주요 브리핑 및 뉴스링크
    briefTasks = []
    mainbrief = bs.findAll('a',{'href' : re.compile('\/tcmBoardView\.do\?contSeq=[0-9]*')})
    for brf in mainbrief:
        container = []
        container.append(brf.text)
        container.append(covidNotice + brf['href'])
        briefTasks.append(container)
    print(briefTasks)

    # 통계수치
    statNum = []
    # 전일대비 수치
    beforeNum = []
    for num in range(7):
        statNum.append(statisticalNumbers[num].text)
    for num in range(4):
        beforeNum.append(beforedayNumbers[num].text.split('(')[-1].split(')')[0])

    totalPeopletoInt = statNum[0].split(')')[-1].split(',')
    totalPeopleDeathtoInt = statNum[3].split(')')[-1].split(',')
    tpInt = ''.join(totalPeopletoInt)
    tpDInt = ''.join(totalPeopleDeathtoInt)
    lethatRate = round((int(tpDInt) / int(tpInt)) * 100, 2)
    
    embed = discord.Embed(title="Covid-19 Virus Korea Status", description="",color=0x5CD1E5)
    embed.add_field(name="Data source : Ministry of Health and Welfare of Korea", value="http://ncov.mohw.go.kr/index.jsp", inline=False)
    embed.add_field(name="Latest data refred time",value="해당 자료는 " + latestupdateTime[0] + "월 " + latestupdateTime[1] + "일 "+latestupdateTime[2] +" 자료입니다.", inline=False)
    embed.add_field(name="확진환자(누적)", value=statNum[0].split(')')[-1]+"("+beforeNum[0]+")",inline=True)
    embed.add_field(name="완치환자(격리해제)", value=statNum[1] + "(" + beforeNum[1] + ")", inline=True)
    embed.add_field(name="치료중(격리 중)", value=statNum[2] + "(" + beforeNum[2] + ")", inline=True)
    embed.add_field(name="사망", value=statNum[3] + "(" + beforeNum[3] + ")", inline=True)
    embed.add_field(name="누적확진률", value=statNum[6], inline=True)
    embed.add_field(name="치사율", value=str(lethatRate) + " %",inline=True)
    embed.add_field(name="- 최신 브리핑 1 : " + briefTasks[0][0],value="Link : " + briefTasks[0][1],inline=False)
    embed.add_field(name="- 최신 브리핑 2 : " + briefTasks[1][0], value="Link : " + briefTasks[1][1], inline=False)
    embed.set_thumbnail(url="https://wikis.krsocsci.org/images/7/79/%EB%8C%80%ED%95%9C%EC%99%95%EA%B5%AD_%ED%83%9C%EA%B7%B9%EA%B8%B0.jpg")
    embed.set_footer(text='Helped by Hoplin.')
    await ctx.send("Covid-19 Virus Korea Status", embed=embed)
    
@bot.command()
async def 영화(ctx):
    movie_rank()
    wb = xl.load_workbook('경로/MovieRank.xlsx')
    sheet1 = wb["영화 순위"]
    rank, title, story, link, poster = [], [], [], [], []

    for cell in sheet1['A']: # 랭크
        rank.append(cell.value)
    for cell in sheet1['B']: # 타이틀
        title.append(cell.value)
    for cell in sheet1['C']: # 스토리
        story.append(cell.value)
    for cell in sheet1['D']: # 링크
        link.append(cell.value)
    for cell in sheet1['E']: # 포스터
        poster.append(cell.value)
    wb.close()
    
    embed = discord.Embed(title="박스오피스 영화 순위", description="현재 박스오피스 1위 ~ 10위 영화", color=0x00ffff)
    embed.add_field(name="===========================================================", value="Link : https://movie.daum.net/ranking/boxoffice/weekly", inline=False)
    embed.add_field(name=rank[1], value=title[1]+'\nㅤ', inline=False)
    #embed.add_field(name="==================================================", value="ㅤ", inline=False)
    embed.add_field(name=rank[2], value=title[2], inline=True)
    embed.add_field(name=rank[3], value=title[3], inline=True)
    embed.add_field(name=rank[4], value=title[4], inline=True)
    embed.add_field(name=rank[5], value=title[5], inline=True)
    embed.add_field(name=rank[6], value=title[6], inline=True)
    embed.add_field(name=rank[7], value=title[7], inline=True)
    embed.add_field(name=rank[8], value=title[8], inline=True)
    embed.add_field(name=rank[9], value=title[9], inline=True)
    embed.add_field(name=rank[10], value=title[10], inline=True)
    embed.add_field(name="===========================================================", value="ㅤ", inline=False)
    embed.add_field(name='현재 1위 영화', value="영화 예매 -----> {}".format(link[1]), inline=False)
    embed.set_image(url=poster[1])
    await ctx.send(embed=embed)

@bot.command()
async def 영화정보(ctx, *, text):
    text = int(text)
    movie_rank()
    
    wb = xl.load_workbook('경로/MovieRank.xlsx')
    sheet1 = wb["영화 순위"]
    rank, title, story, link, poster = [], [], [], [], []

    for cell in sheet1['A']:
        rank.append(cell.value)
    for cell in sheet1['B']:
        title.append(cell.value)
    for cell in sheet1['C']:
        story.append(cell.value)
    for cell in sheet1['D']:
        link.append(cell.value)
    for cell in sheet1['E']:
        poster.append(cell.value)
    wb.close()
        
    movie_info(link[text])
    wb = xl.load_workbook('경로/MovieRank.xlsx')
    sheet2 = wb["영화 정보"]
    info_name, info = [], []

    for cell in sheet2['A']:
        info_name.append(cell.value)
    for cell in sheet2['B']:
        info.append(cell.value)
    wb.close()

    embed = discord.Embed(title=rank[text]+" 영화 정보", description='', color=0x00ffff)
    embed.add_field(name=title[text], value=story[text][:350]+' ... \n(더 많은 정보를 아래 링크를 통해 확인하세요!)', inline=False)
    embed.add_field(name="================================================", value="ㅤ", inline=False)
    embed.add_field(name=info_name[0], value=info[0], inline=True)
    embed.add_field(name=info_name[1], value=info[1], inline=True)
    embed.add_field(name=info_name[2], value=info[2], inline=True)
    embed.add_field(name=info_name[3], value=info[3], inline=True)
    embed.add_field(name=info_name[4], value=info[4], inline=True)
    embed.add_field(name=info_name[5], value=info[5], inline=True)
    embed.add_field(name=info_name[6], value=info[6], inline=True)
    embed.add_field(name="================================================", value="ㅤ", inline=False)
    embed.add_field(name='현재 영화', value="영화 예매하러 가기 -----> {}".format(link[text]), inline=False)
    embed.set_thumbnail(url=poster[text])
    embed.set_image(url=poster[text])
    await ctx.send(embed=embed)
    

bot.run('사용중인 봇 토큰')
