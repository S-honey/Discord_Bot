from bs4 import BeautifulSoup
import requests     # 웹사이트에 접속
import openpyxl as xl
import pandas as pd
import random
import discord
from discord.ext import commands

# 오늘의 운세 크롤링 후 엑셀파일로 저장
def crawl_daily_fortune():
    url = 'https://unse.daily.co.kr/?p=zodiac'
    res = requests.get(url)
    soup = BeautifulSoup(res.text, 'html.parser')

    # 1. 내가 찾고자 하는 데이터를 포함하는 가장 근접한 부모
    parent = soup.select_one('#card')

    # 2. 부모 밑에 있는 자식 요소
    children = parent.select('ul')      # select: 여러개의 요소 / select_one: 가장 첫번째 요소

    wb = xl.Workbook()
    children = parent.select('ul')      # select: 여러개의 요소 / select_one: 가장 첫번째 요소
    for i in children:
        lis = i.select('li')
        for n, li in enumerate(lis):
            if n == 0:
                animal = li.select_one('div > b').get_text()    # 태그 중 텍스트 추출
                total = li.select_one('div > p').get_text()
                #print(animal, total)
                #print()
                ws = wb.create_sheet(animal+'띠')    # animal 이름으로 시트 생성
                ws.append([total])    
            else:
                year = li.select_one('span').get_text()
                luck = li.select_one('p').get_text()
                #print(year, luck)
                #print()
                ws.append([year, luck])
    wb.save('fortune.xlsx')

# 크롤링 코드에서 저장한 엑셀파일 출력
def daily_fortune(animal):
    wb = xl.load_workbook('C:/discordbot/fortune.xlsx') 
    sheet = wb[animal]

    global one_line
    global year_line

    one_line = []
    year_line = []

    # A1 데이터를 가져오는 방법 
    one_line = sheet['A1'].value

    # A1 부터, B1 데이터까지 가져오기 
    for data in sheet['A4':'B6']:
        for cell in data:
            year_line.append(cell.value)

    wb.close()




# 디스코드 봇 실행 코드

game = discord.Game("뭔가")
bot = commands.Bot(command_prefix='ㄹㅇ ', status=discord.Status.online, activity=game)

@bot.event
async def on_ready():
    print('사용중인 봇 :', bot.user.name)
    print('connection was succesful')
    await bot.change_presence(status=discord.Status.online, activity=game)

@bot.command(aliases=['안녕', 'Hi', '안녕하세요', '안녕하십니까'])
async def hello(ctx):
    random_num = random.randrange(1,15)
    if random_num >= 4:
        await ctx.send(f'{ctx.author.mention}님 안녕하세요!')
    elif random_num < 4: # 낮은 확률로 다른 인사
        await ctx.send(f'{ctx.author.mention}님 잘 지내셨나요?')

@bot.command()
async def 도움(ctx):
    embed = discord.Embed(title='도움말', description='제작 by않되어떡케', color = 0x00ff00)
    embed.add_field(name='1. 인사', value='ㄹㅇ 안녕', inline=False)
    embed.add_field(name='2. 운세 [~띠]', value='ㄹㅇ 오늘의[~띠] 운세', inline=False)
    await ctx.send(embed=embed)

@bot.command()
async def 운세(ctx, *, text):
    crawl_daily_fortune()
    fortune = daily_fortune(text)
    embed = discord.Embed(title=text, description=one_line, color=0x00ff55)
    embed.add_field(name=year_line[0], value=year_line[1], inline=False)
    embed.add_field(name=year_line[2], value=year_line[3], inline=False)
    embed.add_field(name=year_line[4], value=year_line[5], inline=False)
    await ctx.send(embed=embed)

bot.run('Your Token')

